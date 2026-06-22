"""Polished, multi-sheet Excel report — the headline deliverable.

Sheets: Summary, Priorities, Hosts, Open Ports, Services, NSE/CVE, Appendix.
Branded header, severity colour coding, AutoFilter, freeze panes, print-ready.
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..core import paths
from ..core.models import ScanResult, SEVERITY_ORDER

# Brand palette (matches the GUI theme)
CHARCOAL = "FF0C1118"
SLATE = "FF161E29"
ACCENT = "FF48F3D2"
ACCENT_DK = "FF0FB497"
TEXT = "FFE7EDF5"

SEV_FILL = {
    "CRITICAL": "FFFF5C7A",
    "HIGH": "FFFF9F43",
    "MEDIUM": "FFFFD166",
    "LOW": "FF62D5B8",
    "INFO": "FFD7DEE8",
}

_THIN = Side(style="thin", color="FFD7DEE8")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _hdr_font():
    return Font(name="Segoe UI", bold=True, color="FF5A6675", size=10)


def _sev_font(sev):
    dark = sev in ("CRITICAL", "HIGH")
    return Font(name="Segoe UI", bold=True, color="FFFFFFFF" if dark else "FF1B2533")


def save_xlsx(result: ScanResult, path: str, t=None) -> str:
    t = t or (lambda k: k)
    wb = Workbook()

    _sheet_summary(wb.active, result, t)
    _sheet_priorities(wb.create_sheet(t("tab.priorities")), result, t)
    _sheet_hosts(wb.create_sheet(t("common.hosts")), result, t)
    _sheet_ports(wb.create_sheet("Open Ports"), result, t)
    _sheet_services(wb.create_sheet(t("col.service")), result, t)
    _sheet_nse(wb.create_sheet("NSE-CVE"), result, t)
    _sheet_appendix(wb.create_sheet("Appendix"), result, t)

    wb.properties.title = "Abilithic Scan Report"
    wb.properties.creator = "Abilithic Scan"
    wb.save(path)
    return path


# --------------------------------------------------------------------------- #
def _brand_header(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ncols, 4))
    c = ws.cell(row=1, column=1, value=f"  🛰  {title}")
    c.fill = PatternFill("solid", fgColor=CHARCOAL)
    c.font = Font(name="Segoe UI", bold=True, size=16, color="FF48F3D2")
    c.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(ncols, 4))
    s = ws.cell(row=2, column=1, value="  " + subtitle)
    s.fill = PatternFill("solid", fgColor=SLATE)
    s.font = Font(name="Segoe UI", size=9, color="FFB7C2D2")
    ws.row_dimensions[2].height = 18
    # try to drop the logo in
    try:
        from openpyxl.drawing.image import Image as XLImage
        logo = paths.resource_path("assets/abilithic-icon-256.png")
        if os.path.exists(logo):
            img = XLImage(logo)
            img.width = img.height = 28
            ws.add_image(img, "A1")
    except Exception:
        pass


def _table_header(ws, row, headers):
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = _hdr_font()
        c.fill = PatternFill("solid", fgColor="FFEEF1F6")
        c.border = _BORDER
        c.alignment = Alignment(vertical="center")


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _sev_cell(ws, row, col, sev, t):
    c = ws.cell(row=row, column=col, value=t(f"severity.{sev}"))
    c.fill = PatternFill("solid", fgColor=SEV_FILL.get(sev, SEV_FILL["INFO"]))
    c.font = _sev_font(sev)
    c.alignment = Alignment(horizontal="center")
    c.border = _BORDER
    return c


# --------------------------------------------------------------------------- #
def _sheet_summary(ws, result, t):
    ws.title = "Summary"
    m = result.scan_meta
    _brand_header(ws, "Abilithic Scan — Report", t("app.subtitle"), 4)
    s = result.summary or {}
    rows = [
        ("Target / Command", m.command),
        ("Profile", m.profile),
        ("Nmap", m.nmap_version),
        ("Started", m.started_at),
        ("Finished", m.finished_at),
        ("Duration (s)", m.duration_s),
        (t("common.hosts"), f"{s.get('hosts_up', 0)} {t('common.up')} / {s.get('hosts_down', 0)} {t('common.down')}"),
        (t("col.open_ports"), s.get("open_ports", 0)),
    ]
    r = 4
    for label, val in rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, color="FF5A6675")
        ws.cell(row=r, column=2, value=str(val))
        r += 1

    r += 1
    ws.cell(row=r, column=1, value=t("col.criticality")).font = Font(bold=True, size=12)
    r += 1
    by = s.get("by_severity", {})
    for sev in reversed(SEVERITY_ORDER):
        _sev_cell(ws, r, 1, sev, t)
        ws.cell(row=r, column=2, value=by.get(sev, 0)).alignment = Alignment(horizontal="center")
        r += 1
    _autosize(ws, [26, 70])
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"


def _sheet_priorities(ws, result, t):
    _brand_header(ws, t("tab.priorities"), t("hint.tab.priorities"), 5)
    headers = ["#", t("col.criticality"), t("col.host"), t("col.problem"), t("col.advice")]
    _table_header(ws, 4, headers)
    r = 5
    if not result.priorities:
        ws.cell(row=r, column=2, value=t("common.none"))
    for i, p in enumerate(result.priorities, start=1):
        ws.cell(row=r, column=1, value=i).border = _BORDER
        _sev_cell(ws, r, 2, p.severity, t)
        ws.cell(row=r, column=3, value=p.host).border = _BORDER
        ws.cell(row=r, column=4, value=p.title).border = _BORDER
        adv = ws.cell(row=r, column=5, value=p.advice)
        adv.alignment = Alignment(wrap_text=True, vertical="top")
        adv.border = _BORDER
        r += 1
    ws.freeze_panes = "A5"
    if r > 5:
        ws.auto_filter.ref = f"A4:E{r-1}"
    _autosize(ws, [5, 14, 24, 34, 60])
    ws.sheet_view.showGridLines = False


def _sheet_hosts(ws, result, t):
    _brand_header(ws, t("common.hosts"), "", 7)
    headers = [t("col.host"), "Hostname", t("col.os"), t("col.exposure"),
               t("col.open_ports"), t("col.risk")]
    _table_header(ws, 4, headers)
    r = 5
    for h in result.hosts:
        ws.cell(row=r, column=1, value=h.address).border = _BORDER
        ws.cell(row=r, column=2, value=h.primary_name()).border = _BORDER
        osname = h.os_matches[0].name if h.os_matches else ""
        ws.cell(row=r, column=3, value=osname).border = _BORDER
        ws.cell(row=r, column=4, value=t(f"common.{h.exposure}") if h.exposure in ("public", "internal") else h.exposure).border = _BORDER
        ws.cell(row=r, column=5, value=len(h.open_ports())).border = _BORDER
        _sev_cell(ws, r, 6, h.risk_grade, t)
        r += 1
    ws.freeze_panes = "A5"
    if r > 5:
        ws.auto_filter.ref = f"A4:F{r-1}"
    _autosize(ws, [18, 26, 30, 12, 12, 14])
    ws.sheet_view.showGridLines = False


def _sheet_ports(ws, result, t):
    _brand_header(ws, "Open Ports", t("hint.tab.ports"), 8)
    headers = [t("col.host"), t("col.port"), t("col.proto"), t("col.service"),
               t("col.version"), t("col.criticality"), t("col.why")]
    _table_header(ws, 4, headers)
    r = 5
    for h in result.hosts:
        for p in h.open_ports():
            ws.cell(row=r, column=1, value=h.address).border = _BORDER
            ws.cell(row=r, column=2, value=p.portid).border = _BORDER
            ws.cell(row=r, column=3, value=p.protocol).border = _BORDER
            ws.cell(row=r, column=4, value=p.service.name).border = _BORDER
            ws.cell(row=r, column=5, value=p.service.label()).border = _BORDER
            _sev_cell(ws, r, 6, p.criticality.severity, t)
            why = "; ".join(c.reason for c in p.criticality.contributors)
            wc = ws.cell(row=r, column=7, value=why)
            wc.alignment = Alignment(wrap_text=True, vertical="top")
            wc.border = _BORDER
            r += 1
    ws.freeze_panes = "A5"
    if r > 5:
        ws.auto_filter.ref = f"A4:G{r-1}"
    _autosize(ws, [18, 8, 9, 18, 26, 14, 60])
    ws.sheet_view.showGridLines = False


def _sheet_services(ws, result, t):
    _brand_header(ws, t("col.service"), "", 3)
    _table_header(ws, 4, [t("col.service"), t("col.version"), t("common.hosts")])
    agg = {}
    for h in result.hosts:
        for p in h.open_ports():
            key = (p.service.name, p.service.label())
            agg[key] = agg.get(key, 0) + 1
    r = 5
    for (name, ver), n in sorted(agg.items(), key=lambda x: -x[1]):
        ws.cell(row=r, column=1, value=name).border = _BORDER
        ws.cell(row=r, column=2, value=ver).border = _BORDER
        ws.cell(row=r, column=3, value=n).border = _BORDER
        r += 1
    ws.freeze_panes = "A5"
    if r > 5:
        ws.auto_filter.ref = f"A4:C{r-1}"
    _autosize(ws, [24, 30, 10])
    ws.sheet_view.showGridLines = False


def _sheet_nse(ws, result, t):
    _brand_header(ws, "NSE / CVE", t("hint.tab.nse"), 6)
    _table_header(ws, 4, [t("col.host"), t("col.port"), "Script",
                          t("col.cve"), "KEV", "Output"])
    r = 5
    for h in result.hosts:
        for p in h.open_ports():
            for sc in p.scripts:
                ws.cell(row=r, column=1, value=h.address).border = _BORDER
                ws.cell(row=r, column=2, value=p.portid).border = _BORDER
                ws.cell(row=r, column=3, value=sc.id).border = _BORDER
                ws.cell(row=r, column=4, value=", ".join(sc.cve)).border = _BORDER
                ws.cell(row=r, column=5, value=t("common.yes") if sc.kev else "").border = _BORDER
                oc = ws.cell(row=r, column=6, value=sc.output[:1500])
                oc.alignment = Alignment(wrap_text=True, vertical="top")
                oc.border = _BORDER
                r += 1
    if r == 5:
        ws.cell(row=r, column=1, value=t("common.none"))
    ws.freeze_panes = "A5"
    if r > 5:
        ws.auto_filter.ref = f"A4:F{r-1}"
    _autosize(ws, [16, 8, 24, 20, 8, 80])
    ws.sheet_view.showGridLines = False


def _sheet_appendix(ws, result, t):
    _brand_header(ws, "Appendix", "", 2)
    m = result.scan_meta
    rows = [
        ("App version", m.app_version),
        ("Nmap version", m.nmap_version),
        ("Command", m.command),
        ("Arguments", " ".join(m.scan_args)),
        ("Locale", m.locale),
        ("Elevated", str(m.elevated)),
        ("Cancelled", str(m.was_cancelled)),
        ("Error", m.error or "-"),
    ]
    r = 4
    for k, v in rows:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, color="FF5A6675")
        c = ws.cell(row=r, column=2, value=v)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    _autosize(ws, [18, 100])
    ws.sheet_view.showGridLines = False
